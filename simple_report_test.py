#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
簡化的股票報告測試 - 用於調試
"""

import sys
import os
import sqlite3
import pandas as pd
from datetime import datetime

# 添加專案根目錄到 Python 路徑
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from config import Config

def test_basic_data(stock_id):
    """測試基本資料獲取"""
    print(f"測試股票 {stock_id} 的基本資料...")
    
    try:
        conn = sqlite3.connect(Config.DATABASE_PATH)
        cursor = conn.cursor()
        
        # 檢查股票是否存在
        cursor.execute("SELECT * FROM stocks WHERE stock_id = ?", (stock_id,))
        stock_info = cursor.fetchone()
        
        if not stock_info:
            print(f"❌ 股票 {stock_id} 不存在")
            return False
        
        print(f"✅ 股票基本資料: {stock_info[1]} ({stock_info[0]})")
        
        # 檢查股價資料
        cursor.execute("SELECT COUNT(*) FROM stock_prices WHERE stock_id = ?", (stock_id,))
        price_count = cursor.fetchone()[0]
        print(f"📊 股價資料: {price_count} 筆")
        
        # 檢查月營收資料
        cursor.execute("SELECT COUNT(*) FROM monthly_revenues WHERE stock_id = ?", (stock_id,))
        revenue_count = cursor.fetchone()[0]
        print(f"📈 月營收資料: {revenue_count} 筆")
        
        # 檢查財務報表
        cursor.execute("SELECT COUNT(*) FROM financial_statements WHERE stock_id = ?", (stock_id,))
        financial_count = cursor.fetchone()[0]
        print(f"📋 財務報表: {financial_count} 筆")
        
        # 檢查股利政策
        cursor.execute("SELECT COUNT(*) FROM dividend_policies WHERE stock_id = ?", (stock_id,))
        dividend_count = cursor.fetchone()[0]
        print(f"💰 股利政策: {dividend_count} 筆")
        
        # 檢查潛力分析
        cursor.execute("SELECT COUNT(*) FROM stock_scores WHERE stock_id = ?", (stock_id,))
        score_count = cursor.fetchone()[0]
        print(f"🎯 潛力分析: {score_count} 筆")
        
        conn.close()
        return True
        
    except Exception as e:
        print(f"❌ 測試失敗: {e}")
        return False

def test_excel_creation():
    """測試Excel創建"""
    print("測試Excel創建功能...")
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        
        # 創建簡單的Excel檔案
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "測試"
        
        # 添加測試資料
        ws['A1'] = "測試標題"
        ws['A2'] = "測試內容"
        ws['B2'] = "123.45"
        
        # 設定樣式
        ws['A1'].font = Font(bold=True)
        
        # 儲存檔案
        test_filename = "excel_test.xlsx"
        wb.save(test_filename)
        
        print(f"✅ Excel測試檔案創建成功: {test_filename}")
        
        # 清理測試檔案
        if os.path.exists(test_filename):
            os.remove(test_filename)
            print("🗑️ 測試檔案已清理")
        
        return True
        
    except ImportError as e:
        print(f"❌ 缺少必要套件: {e}")
        print("請執行: pip install openpyxl")
        return False
    except Exception as e:
        print(f"❌ Excel測試失敗: {e}")
        return False

def create_simple_report(stock_id):
    """創建簡化報告"""
    print(f"創建 {stock_id} 的簡化報告...")
    
    try:
        import openpyxl
        
        # 獲取基本資料
        conn = sqlite3.connect(Config.DATABASE_PATH)
        cursor = conn.cursor()
        
        cursor.execute("SELECT stock_id, stock_name, market FROM stocks WHERE stock_id = ?", (stock_id,))
        stock_info = cursor.fetchone()
        
        if not stock_info:
            print(f"❌ 股票 {stock_id} 不存在")
            return False
        
        # 創建Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "基本資訊"
        
        # 添加基本資訊
        ws['A1'] = "股票代號"
        ws['B1'] = stock_info[0]
        ws['A2'] = "股票名稱"
        ws['B2'] = stock_info[1]
        ws['A3'] = "所屬市場"
        ws['B3'] = "上市" if stock_info[2] == 'TWSE' else "上櫃"
        
        # 設定樣式
        for row in range(1, 4):
            ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
        
        # 儲存檔案
        filename = f"{stock_id}_簡化報告_{datetime.now().strftime('%Y%m%d')}.xlsx"
        wb.save(filename)
        
        print(f"✅ 簡化報告創建成功: {filename}")
        conn.close()
        return True
        
    except Exception as e:
        print(f"❌ 簡化報告創建失敗: {e}")
        return False

def main():
    """主函數"""
    print("=" * 50)
    print("📊 簡化股票報告測試")
    print("=" * 50)
    
    # 測試股票代號
    test_stocks = ['2330', '0050', '2454']
    
    for stock_id in test_stocks:
        print(f"\n🧪 測試股票: {stock_id}")
        
        # 測試基本資料
        if test_basic_data(stock_id):
            print(f"✅ {stock_id} 基本資料測試通過")
            
            # 測試Excel創建
            if test_excel_creation():
                print("✅ Excel功能測試通過")
                
                # 創建簡化報告
                if create_simple_report(stock_id):
                    print(f"🎉 {stock_id} 簡化報告創建成功！")
                    break  # 成功一個就停止
                else:
                    print(f"❌ {stock_id} 簡化報告創建失敗")
            else:
                print("❌ Excel功能測試失敗")
        else:
            print(f"❌ {stock_id} 基本資料測試失敗")
    
    print("\n" + "=" * 50)
    print("測試完成")
    print("=" * 50)

if __name__ == "__main__":
    main()
