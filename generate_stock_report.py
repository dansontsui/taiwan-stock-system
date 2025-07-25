#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
股票財務分析報告生成器
輸入股票代號，生成完整的Excel財務分析報告
"""

import sys
import os
import sqlite3
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import argparse

# 添加專案根目錄到 Python 路徑
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from config import Config

class StockReportGenerator:
    def __init__(self, stock_id):
        self.stock_id = stock_id.upper()
        self.db_path = Config.DATABASE_PATH
        self.report_data = {}
        self.stock_info = None
        
    def validate_stock_id(self):
        """驗證股票代號是否存在"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM stocks WHERE stock_id = ?", (self.stock_id,))
        self.stock_info = cursor.fetchone()
        conn.close()
        
        if not self.stock_info:
            return False, f"股票代號 {self.stock_id} 不存在於資料庫中"
        
        return True, "股票代號驗證成功"
    
    def get_basic_info(self):
        """獲取基本資訊"""
        if not self.stock_info:
            return None
            
        conn = sqlite3.connect(self.db_path)
        
        # 基本資訊
        basic_info = {
            '股票代號': self.stock_info[0],
            '股票名稱': self.stock_info[1],
            '所屬市場': '上市' if self.stock_info[2] == 'TWSE' else '上櫃',
            '產業別': self.stock_info[3] or '無資料',
            '上市日期': self.stock_info[4] or '無資料',
            '是否為ETF': '是' if self.stock_info[5] else '否'
        }
        
        # 最新股價資訊
        cursor = conn.cursor()
        cursor.execute("""
            SELECT date, close_price, volume, trading_money
            FROM stock_prices 
            WHERE stock_id = ? 
            ORDER BY date DESC 
            LIMIT 1
        """, (self.stock_id,))
        
        latest_price = cursor.fetchone()
        if latest_price:
            basic_info.update({
                '最新交易日': latest_price[0],
                '最新股價': f"{latest_price[1]:.2f}",
                '成交量': f"{latest_price[2]:,}",
                '成交金額': f"{latest_price[3]:,.0f}" if latest_price[3] else '無資料'
            })
        else:
            basic_info.update({
                '最新交易日': '無資料',
                '最新股價': '無資料',
                '成交量': '無資料',
                '成交金額': '無資料'
            })
        
        # 財務比率 - 使用容錯查詢
        try:
            cursor.execute("""
                SELECT pe_ratio, pb_ratio, dividend_yield
                FROM financial_ratios
                WHERE stock_id = ?
                ORDER BY date DESC
                LIMIT 1
            """, (self.stock_id,))

            ratios = cursor.fetchone()
            if ratios:
                basic_info.update({
                    '本益比(PE)': f"{ratios[0]:.2f}" if ratios[0] else '無資料',
                    '股價淨值比(PB)': f"{ratios[1]:.2f}" if ratios[1] else '無資料',
                    '殖利率': f"{ratios[2]:.2f}%" if ratios[2] else '無資料'
                })
            else:
                basic_info.update({
                    '本益比(PE)': '無資料',
                    '股價淨值比(PB)': '無資料',
                    '殖利率': '無資料'
                })
        except Exception:
            # 如果財務比率表不存在或欄位不匹配，設為無資料
            basic_info.update({
                '本益比(PE)': '無資料',
                '股價淨值比(PB)': '無資料',
                '殖利率': '無資料'
            })
        
        conn.close()
        return basic_info
    
    def get_monthly_revenue(self):
        """獲取月營收資料（近24個月）"""
        conn = sqlite3.connect(self.db_path)
        
        query = """
            SELECT revenue_year, revenue_month, revenue, 
                   revenue_growth_mom, revenue_growth_yoy
            FROM monthly_revenues 
            WHERE stock_id = ? 
            ORDER BY revenue_year DESC, revenue_month DESC 
            LIMIT 24
        """
        
        df = pd.read_sql_query(query, conn, params=(self.stock_id,))
        conn.close()
        
        if df.empty:
            return pd.DataFrame(columns=['年月', '營收金額', '月增率(%)', '年增率(%)'])
        
        # 格式化資料
        df['年月'] = df['revenue_year'].astype(str) + '/' + df['revenue_month'].astype(str).str.zfill(2)
        df['營收金額'] = df['revenue'].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else '無資料')
        df['月增率(%)'] = df['revenue_growth_mom'].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else '無資料')
        df['年增率(%)'] = df['revenue_growth_yoy'].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else '無資料')
        
        return df[['年月', '營收金額', '月增率(%)', '年增率(%)']].reset_index(drop=True)
    
    def get_quarterly_financials(self):
        """獲取季度財務資料（近8季）"""
        conn = sqlite3.connect(self.db_path)
        
        # 獲取綜合損益表資料
        query = """
            SELECT date, type, value, origin_name
            FROM financial_statements 
            WHERE stock_id = ? 
            AND type IN ('Revenue', 'GrossProfit', 'OperatingIncome', 'NetIncome', 'EPS')
            ORDER BY date DESC
        """
        
        df = pd.read_sql_query(query, conn, params=(self.stock_id,))
        conn.close()
        
        if df.empty:
            return pd.DataFrame(columns=['季度', '營業收入', '毛利', '營業利益', '稅後淨利', 'EPS'])
        
        # 轉換為透視表格式
        pivot_df = df.pivot_table(index='date', columns='type', values='value', aggfunc='first')
        pivot_df = pivot_df.head(8)  # 近8季
        
        # 格式化
        result_df = pd.DataFrame()
        result_df['季度'] = pivot_df.index
        result_df['營業收入'] = pivot_df.get('Revenue', pd.Series()).apply(lambda x: f"{x:,.0f}" if pd.notna(x) else '無資料')
        result_df['毛利'] = pivot_df.get('GrossProfit', pd.Series()).apply(lambda x: f"{x:,.0f}" if pd.notna(x) else '無資料')
        result_df['營業利益'] = pivot_df.get('OperatingIncome', pd.Series()).apply(lambda x: f"{x:,.0f}" if pd.notna(x) else '無資料')
        result_df['稅後淨利'] = pivot_df.get('NetIncome', pd.Series()).apply(lambda x: f"{x:,.0f}" if pd.notna(x) else '無資料')
        result_df['EPS'] = pivot_df.get('EPS', pd.Series()).apply(lambda x: f"{x:.2f}" if pd.notna(x) else '無資料')
        
        return result_df.reset_index(drop=True)
    
    def get_annual_financials(self):
        """獲取年度財務資料（近5年）"""
        conn = sqlite3.connect(self.db_path)
        
        # 獲取資產負債表和財務比率
        query_balance = """
            SELECT date, type, value
            FROM balance_sheets 
            WHERE stock_id = ? 
            AND type IN ('TotalAssets', 'TotalEquity', 'TotalLiabilities')
            ORDER BY date DESC
        """
        
        query_ratios = """
            SELECT date, roe, roa, gross_margin, operating_margin, net_margin
            FROM financial_ratios
            WHERE stock_id = ?
            ORDER BY date DESC
            LIMIT 5
        """

        try:
            df_balance = pd.read_sql_query(query_balance, conn, params=(self.stock_id,))
        except Exception:
            df_balance = pd.DataFrame()

        try:
            df_ratios = pd.read_sql_query(query_ratios, conn, params=(self.stock_id,))
        except Exception:
            df_ratios = pd.DataFrame()
        conn.close()
        
        # 處理資產負債表資料
        if not df_balance.empty:
            balance_pivot = df_balance.pivot_table(index='date', columns='type', values='value', aggfunc='first')
            balance_pivot = balance_pivot.head(5)
        else:
            balance_pivot = pd.DataFrame()
        
        # 合併資料
        result_df = pd.DataFrame()
        
        if not df_ratios.empty:
            result_df['年度'] = df_ratios['date']
            result_df['ROE(%)'] = df_ratios['roe'].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else '無資料')
            result_df['ROA(%)'] = df_ratios['roa'].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else '無資料')
            result_df['毛利率(%)'] = df_ratios['gross_margin'].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else '無資料')
            result_df['營業利益率(%)'] = df_ratios['operating_margin'].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else '無資料')
            result_df['淨利率(%)'] = df_ratios['net_margin'].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else '無資料')
        
        if not balance_pivot.empty:
            if 'TotalAssets' in balance_pivot.columns:
                result_df['總資產'] = balance_pivot['TotalAssets'].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else '無資料')
            if 'TotalEquity' in balance_pivot.columns:
                result_df['股東權益'] = balance_pivot['TotalEquity'].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else '無資料')
            if 'TotalLiabilities' in balance_pivot.columns:
                result_df['總負債'] = balance_pivot['TotalLiabilities'].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else '無資料')
        
        if result_df.empty:
            return pd.DataFrame(columns=['年度', 'ROE(%)', 'ROA(%)', '毛利率(%)', '營業利益率(%)', '淨利率(%)', '總資產', '股東權益', '總負債'])
        
        return result_df.reset_index(drop=True)
    
    def get_dividend_policy(self):
        """獲取股利政策（近5年）"""
        conn = sqlite3.connect(self.db_path)

        # 先檢查表結構，使用實際的欄位名稱
        query = """
            SELECT date, year,
                   cash_earnings_distribution, stock_earnings_distribution,
                   cash_ex_dividend_trading_date, stock_ex_dividend_trading_date
            FROM dividend_policies
            WHERE stock_id = ?
            ORDER BY date DESC
            LIMIT 5
        """

        try:
            df = pd.read_sql_query(query, conn, params=(self.stock_id,))
        except Exception as e:
            # 如果上述查詢失敗，嘗試簡化查詢
            try:
                query_simple = """
                    SELECT date, year, cash_earnings_distribution, stock_earnings_distribution
                    FROM dividend_policies
                    WHERE stock_id = ?
                    ORDER BY date DESC
                    LIMIT 5
                """
                df = pd.read_sql_query(query_simple, conn, params=(self.stock_id,))
            except:
                conn.close()
                return pd.DataFrame(columns=['年度', '現金股利', '股票股利', '除息日', '除權日'])

        conn.close()

        if df.empty:
            return pd.DataFrame(columns=['年度', '現金股利', '股票股利', '除息日', '除權日'])

        # 格式化資料
        result_df = pd.DataFrame()
        result_df['年度'] = df['year'] if 'year' in df.columns else df['date']
        result_df['現金股利'] = df['cash_earnings_distribution'].apply(lambda x: f"{x:.2f}" if pd.notna(x) else '無資料')
        result_df['股票股利'] = df['stock_earnings_distribution'].apply(lambda x: f"{x:.2f}" if pd.notna(x) else '無資料')
        result_df['除息日'] = df.get('cash_ex_dividend_trading_date', pd.Series()).fillna('無資料')
        result_df['除權日'] = df.get('stock_ex_dividend_trading_date', pd.Series()).fillna('無資料')

        return result_df.reset_index(drop=True)

    def get_cash_flow_data(self):
        """獲取現金流量表資料（近8季）"""
        conn = sqlite3.connect(self.db_path)

        try:
            query = """
                SELECT date, type, value, origin_name
                FROM cash_flow_statements
                WHERE stock_id = ?
                ORDER BY date DESC
                LIMIT 32
            """

            df = pd.read_sql_query(query, conn, params=(self.stock_id,))
        except Exception as e:
            # 如果表不存在或查詢失敗，返回空DataFrame
            conn.close()
            return pd.DataFrame(columns=['季度', '營業現金流', '投資現金流', '融資現金流', '自由現金流'])

        conn.close()

        if df.empty:
            return pd.DataFrame(columns=['季度', '營業現金流', '投資現金流', '融資現金流', '自由現金流'])

        # 轉換為季度格式並透視
        df['quarter'] = df['date'].apply(lambda x: f"{x[:4]}Q{(int(x[5:7])-1)//3+1}")

        # 分類現金流量類型
        operating_cf = df[df['type'].str.contains('營業|Operating', case=False, na=False)]
        investing_cf = df[df['type'].str.contains('投資|Investing', case=False, na=False)]
        financing_cf = df[df['type'].str.contains('融資|Financing', case=False, na=False)]

        # 按季度匯總
        result_data = []
        quarters = df['quarter'].unique()[:8]  # 近8季

        for quarter in quarters:
            operating = operating_cf[operating_cf['quarter'] == quarter]['value'].sum()
            investing = investing_cf[investing_cf['quarter'] == quarter]['value'].sum()
            financing = financing_cf[financing_cf['quarter'] == quarter]['value'].sum()
            free_cf = operating + investing  # 自由現金流 = 營業現金流 + 投資現金流

            result_data.append({
                '季度': quarter,
                '營業現金流': f"{operating:,.0f}" if operating != 0 else '無資料',
                '投資現金流': f"{investing:,.0f}" if investing != 0 else '無資料',
                '融資現金流': f"{financing:,.0f}" if financing != 0 else '無資料',
                '自由現金流': f"{free_cf:,.0f}" if free_cf != 0 else '無資料'
            })

        return pd.DataFrame(result_data)

    def get_dividend_results(self):
        """獲取除權除息結果（近5年）"""
        conn = sqlite3.connect(self.db_path)

        try:
            query = """
                SELECT date, before_price, after_price,
                       stock_and_cache_dividend, stock_or_cache_dividend,
                       max_price, min_price, open_price, reference_price
                FROM dividend_results
                WHERE stock_id = ?
                ORDER BY date DESC
                LIMIT 20
            """

            df = pd.read_sql_query(query, conn, params=(self.stock_id,))
        except Exception as e:
            # 如果表不存在或查詢失敗，返回空DataFrame
            conn.close()
            return pd.DataFrame(columns=['除權息日', '除權息前價格', '除權息後價格', '股利金額', '填權息表現'])

        conn.close()

        if df.empty:
            return pd.DataFrame(columns=['除權息日', '除權息前價格', '除權息後價格', '股利金額', '填權息表現'])

        # 計算填權息表現
        df['填權息表現'] = df.apply(lambda row:
            f"{((row['after_price'] - row['reference_price']) / row['reference_price'] * 100):.2f}%"
            if pd.notna(row['after_price']) and pd.notna(row['reference_price']) and row['reference_price'] != 0
            else '無資料', axis=1)

        result_df = pd.DataFrame()
        result_df['除權息日'] = df['date']
        result_df['除權息前價格'] = df['before_price'].apply(lambda x: f"{x:.2f}" if pd.notna(x) else '無資料')
        result_df['除權息後價格'] = df['after_price'].apply(lambda x: f"{x:.2f}" if pd.notna(x) else '無資料')
        result_df['股利金額'] = df['stock_and_cache_dividend'].apply(lambda x: f"{x:.2f}" if pd.notna(x) else '無資料')
        result_df['填權息表現'] = df['填權息表現']

        return result_df.reset_index(drop=True)

    def get_stock_price_analysis(self):
        """獲取股價分析（近1年）"""
        conn = sqlite3.connect(self.db_path)

        # 使用容錯查詢，嘗試不同的欄位名稱
        try:
            # 先嘗試標準欄位名稱
            query = """
                SELECT date, open_price, high_price, low_price, close_price, volume
                FROM stock_prices
                WHERE stock_id = ? AND date >= date('now', '-1 year')
                ORDER BY date DESC
            """
            df = pd.read_sql_query(query, conn, params=(self.stock_id,))
        except:
            try:
                # 嘗試簡化的欄位名稱
                query = """
                    SELECT date, open, high, low, close, volume
                    FROM stock_prices
                    WHERE stock_id = ? AND date >= date('now', '-1 year')
                    ORDER BY date DESC
                """
                df = pd.read_sql_query(query, conn, params=(self.stock_id,))
                # 重新命名欄位以保持一致性
                df = df.rename(columns={
                    'open': 'open_price',
                    'high': 'high_price',
                    'low': 'low_price',
                    'close': 'close_price'
                })
            except:
                # 如果都失敗，返回無資料
                conn.close()
                return {
                    '當前股價': '無資料',
                    '52週最高': '無資料',
                    '52週最低': '無資料',
                    '平均成交量': '無資料',
                    '股價波動率': '無資料',
                    '近期趨勢': '無資料'
                }

        conn.close()

        if df.empty:
            return {
                '當前股價': '無資料',
                '52週最高': '無資料',
                '52週最低': '無資料',
                '平均成交量': '無資料',
                '股價波動率': '無資料',
                '近期趨勢': '無資料'
            }

        try:
            # 計算各項指標
            current_price = df.iloc[0]['close_price'] if not df.empty else 0
            high_52w = df['high_price'].max()
            low_52w = df['low_price'].min()
            avg_volume = df['volume'].mean() if 'volume' in df.columns else 0

            # 計算波動率（標準差/平均價格）
            volatility = (df['close_price'].std() / df['close_price'].mean() * 100) if len(df) > 1 else 0

            # 判斷趨勢（比較近30天與前30天的平均價格）
            if len(df) >= 60:
                recent_avg = df.head(30)['close_price'].mean()
                previous_avg = df.iloc[30:60]['close_price'].mean()
                trend = "上升" if recent_avg > previous_avg else "下降"
            else:
                trend = "資料不足"

            return {
                '當前股價': f"{current_price:.2f}" if current_price > 0 else '無資料',
                '52週最高': f"{high_52w:.2f}" if pd.notna(high_52w) else '無資料',
                '52週最低': f"{low_52w:.2f}" if pd.notna(low_52w) else '無資料',
                '平均成交量': f"{avg_volume:,.0f}" if avg_volume > 0 else '無資料',
                '股價波動率': f"{volatility:.2f}%" if volatility > 0 else '無資料',
                '近期趨勢': trend
            }
        except Exception as e:
            return {
                '當前股價': '計算錯誤',
                '52週最高': '計算錯誤',
                '52週最低': '計算錯誤',
                '平均成交量': '計算錯誤',
                '股價波動率': '計算錯誤',
                '近期趨勢': '計算錯誤'
            }

    def get_financial_ratios_analysis(self):
        """獲取財務比率分析"""
        conn = sqlite3.connect(self.db_path)

        try:
            query = """
                SELECT date, current_ratio, quick_ratio, debt_ratio,
                       operating_cash_flow, cash_flow_quality
                FROM financial_ratios
                WHERE stock_id = ?
                ORDER BY date DESC
                LIMIT 8
            """

            df = pd.read_sql_query(query, conn, params=(self.stock_id,))
        except Exception as e:
            # 如果表不存在或查詢失敗，返回空DataFrame
            conn.close()
            return pd.DataFrame(columns=['日期', '流動比率', '速動比率', '負債比率', '營業現金流', '現金流量品質'])

        conn.close()

        if df.empty:
            return pd.DataFrame(columns=['日期', '流動比率', '速動比率', '負債比率', '營業現金流', '現金流量品質'])

        result_df = pd.DataFrame()
        result_df['日期'] = df['date']
        result_df['流動比率'] = df['current_ratio'].apply(lambda x: f"{x:.2f}" if pd.notna(x) else '無資料')
        result_df['速動比率'] = df['quick_ratio'].apply(lambda x: f"{x:.2f}" if pd.notna(x) else '無資料')
        result_df['負債比率'] = df['debt_ratio'].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else '無資料')
        result_df['營業現金流'] = df['operating_cash_flow'].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else '無資料')
        result_df['現金流量品質'] = df['cash_flow_quality'].apply(lambda x: f"{x:.2f}" if pd.notna(x) else '無資料')

        return result_df.reset_index(drop=True)

    def get_potential_analysis(self):
        """獲取潛力分析"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # 先嘗試使用完整的欄位名稱
        try:
            query = """
                SELECT total_score, growth_score, profitability_score,
                       stability_score, valuation_score, dividend_score,
                       analysis_date
                FROM stock_scores
                WHERE stock_id = ?
                ORDER BY analysis_date DESC
                LIMIT 1
            """
            cursor.execute(query, (self.stock_id,))
            result = cursor.fetchone()
        except Exception:
            # 如果失敗，嘗試簡化查詢
            try:
                query_simple = """
                    SELECT total_score, growth_score, profitability_score,
                           analysis_date
                    FROM stock_scores
                    WHERE stock_id = ?
                    ORDER BY analysis_date DESC
                    LIMIT 1
                """
                cursor.execute(query_simple, (self.stock_id,))
                result = cursor.fetchone()
                if result:
                    # 補充缺失的欄位
                    result = result + (None, None, None)  # 補充3個None值
            except Exception:
                result = None

        conn.close()

        if not result:
            return {
                '總分': '無資料',
                '評等': '無資料',
                '成長潛力': '無資料',
                '獲利能力': '無資料',
                '穩定性': '無資料',
                '估值合理性': '無資料',
                '配息穩定性': '無資料',
                '分析日期': '無資料'
            }

        # 計算評等
        total_score = result[0]
        if total_score and total_score >= 85:
            grade = 'A+'
        elif total_score and total_score >= 75:
            grade = 'A'
        elif total_score and total_score >= 65:
            grade = 'B+'
        elif total_score and total_score >= 55:
            grade = 'B'
        elif total_score and total_score >= 45:
            grade = 'C+'
        elif total_score and total_score >= 35:
            grade = 'C'
        elif total_score:
            grade = 'D'
        else:
            grade = '無資料'

        # 安全的格式化函數
        def safe_format(value):
            if value is None:
                return '無資料'
            try:
                # 確保是數字類型
                if isinstance(value, (int, float)):
                    return f"{value:.1f}"
                elif isinstance(value, str):
                    try:
                        num_value = float(value)
                        return f"{num_value:.1f}"
                    except ValueError:
                        return '無資料'
                else:
                    return '無資料'
            except:
                return '無資料'

        return {
            '總分': safe_format(total_score),
            '評等': grade,
            '成長潛力': safe_format(result[1]) if len(result) > 1 else '無資料',
            '獲利能力': safe_format(result[2]) if len(result) > 2 else '無資料',
            '穩定性': safe_format(result[3]) if len(result) > 3 else '無資料',
            '估值合理性': safe_format(result[4]) if len(result) > 4 else '無資料',
            '配息穩定性': safe_format(result[5]) if len(result) > 5 else '無資料',
            '分析日期': result[-1] if result[-1] else '無資料'  # 使用最後一個元素作為日期
        }

    def generate_excel_report(self):
        """生成Excel報告"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.chart import LineChart, Reference
        except ImportError:
            print("錯誤: 需要安裝 openpyxl 套件")
            print("請執行: pip install openpyxl")
            return False

        # 驗證股票代號
        is_valid, message = self.validate_stock_id()
        if not is_valid:
            print(f"錯誤: {message}")
            return False

        print(f"開始生成 {self.stock_id} 的財務分析報告...")

        # 收集所有資料
        basic_info = self.get_basic_info()
        monthly_revenue = self.get_monthly_revenue()
        quarterly_financials = self.get_quarterly_financials()
        annual_financials = self.get_annual_financials()
        dividend_policy = self.get_dividend_policy()
        cash_flow_data = self.get_cash_flow_data()
        dividend_results = self.get_dividend_results()
        stock_price_analysis = self.get_stock_price_analysis()
        financial_ratios = self.get_financial_ratios_analysis()
        potential_analysis = self.get_potential_analysis()

        # 創建Excel工作簿
        wb = openpyxl.Workbook()

        # 定義樣式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')

        # 1. 基本資訊工作表
        ws1 = wb.active
        ws1.title = "基本資訊"

        # 添加標題
        ws1['A1'] = f"{basic_info['股票名稱']} ({basic_info['股票代號']}) 基本資訊"
        ws1['A1'].font = Font(size=16, bold=True)
        ws1.merge_cells('A1:B1')

        # 添加基本資訊
        row = 3
        for key, value in basic_info.items():
            ws1[f'A{row}'] = key
            ws1[f'B{row}'] = value
            ws1[f'A{row}'].font = Font(bold=True)
            row += 1

        # 格式化
        for row in ws1.iter_rows(min_row=3, max_row=row-1, min_col=1, max_col=2):
            for cell in row:
                cell.border = border

        # 2. 月營收工作表
        ws2 = wb.create_sheet("月營收")
        ws2['A1'] = "近24個月營收資料"
        ws2['A1'].font = Font(size=14, bold=True)

        if not monthly_revenue.empty:
            # 添加表頭
            for col, header in enumerate(monthly_revenue.columns, 1):
                cell = ws2.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border

            # 添加資料
            for row_idx, row_data in enumerate(monthly_revenue.itertuples(index=False), 4):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = center_alignment
        else:
            ws2['A3'] = "無月營收資料"

        # 3. 季度財務工作表
        ws3 = wb.create_sheet("季度財務")
        ws3['A1'] = "近8季財務資料"
        ws3['A1'].font = Font(size=14, bold=True)

        if not quarterly_financials.empty:
            # 添加表頭
            for col, header in enumerate(quarterly_financials.columns, 1):
                cell = ws3.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border

            # 添加資料
            for row_idx, row_data in enumerate(quarterly_financials.itertuples(index=False), 4):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws3.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = center_alignment
        else:
            ws3['A3'] = "無季度財務資料"

        # 4. 年度財務工作表
        ws4 = wb.create_sheet("年度財務")
        ws4['A1'] = "近5年財務資料"
        ws4['A1'].font = Font(size=14, bold=True)

        if not annual_financials.empty:
            # 添加表頭
            for col, header in enumerate(annual_financials.columns, 1):
                cell = ws4.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border

            # 添加資料
            for row_idx, row_data in enumerate(annual_financials.itertuples(index=False), 4):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws4.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = center_alignment
        else:
            ws4['A3'] = "無年度財務資料"

        # 5. 股利政策工作表
        ws5 = wb.create_sheet("股利政策")
        ws5['A1'] = "近5年股利政策"
        ws5['A1'].font = Font(size=14, bold=True)

        if not dividend_policy.empty:
            # 添加表頭
            for col, header in enumerate(dividend_policy.columns, 1):
                cell = ws5.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border

            # 添加資料
            for row_idx, row_data in enumerate(dividend_policy.itertuples(index=False), 4):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws5.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = center_alignment
        else:
            ws5['A3'] = "無股利政策資料"

        # 6. 現金流量工作表
        ws6 = wb.create_sheet("現金流量")
        ws6['A1'] = "近8季現金流量資料"
        ws6['A1'].font = Font(size=14, bold=True)

        if not cash_flow_data.empty:
            # 添加表頭
            for col, header in enumerate(cash_flow_data.columns, 1):
                cell = ws6.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border

            # 添加資料
            for row_idx, row_data in enumerate(cash_flow_data.itertuples(index=False), 4):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws6.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = center_alignment
        else:
            ws6['A3'] = "無現金流量資料"

        # 7. 除權除息結果工作表
        ws7 = wb.create_sheet("除權除息結果")
        ws7['A1'] = "近5年除權除息結果"
        ws7['A1'].font = Font(size=14, bold=True)

        if not dividend_results.empty:
            # 添加表頭
            for col, header in enumerate(dividend_results.columns, 1):
                cell = ws7.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border

            # 添加資料
            for row_idx, row_data in enumerate(dividend_results.itertuples(index=False), 4):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws7.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = center_alignment
        else:
            ws7['A3'] = "無除權除息結果資料"

        # 8. 股價分析工作表
        ws8 = wb.create_sheet("股價分析")
        ws8['A1'] = "股價技術分析"
        ws8['A1'].font = Font(size=14, bold=True)

        # 添加股價分析資料
        row = 3
        for key, value in stock_price_analysis.items():
            ws8[f'A{row}'] = key
            ws8[f'B{row}'] = value
            ws8[f'A{row}'].font = Font(bold=True)
            row += 1

        # 格式化股價分析
        for row in ws8.iter_rows(min_row=3, max_row=row-1, min_col=1, max_col=2):
            for cell in row:
                cell.border = border

        # 9. 財務比率工作表
        ws9 = wb.create_sheet("財務比率")
        ws9['A1'] = "財務比率分析"
        ws9['A1'].font = Font(size=14, bold=True)

        if not financial_ratios.empty:
            # 添加表頭
            for col, header in enumerate(financial_ratios.columns, 1):
                cell = ws9.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border

            # 添加資料
            for row_idx, row_data in enumerate(financial_ratios.itertuples(index=False), 4):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws9.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = center_alignment
        else:
            ws9['A3'] = "無財務比率資料"

        # 10. 潛力分析工作表
        ws10 = wb.create_sheet("潛力分析")
        ws10['A1'] = "潛力股分析"
        ws10['A1'].font = Font(size=14, bold=True)

        # 添加潛力分析資料
        row = 3
        for key, value in potential_analysis.items():
            ws10[f'A{row}'] = key
            ws10[f'B{row}'] = value
            ws10[f'A{row}'].font = Font(bold=True)
            row += 1

        # 格式化潛力分析
        for row in ws10.iter_rows(min_row=3, max_row=row-1, min_col=1, max_col=2):
            for cell in row:
                cell.border = border

        # 調整欄寬 - 簡化版本避免錯誤
        for ws in [ws1, ws2, ws3, ws4, ws5, ws6, ws7, ws8, ws9, ws10]:
            try:
                # 設定固定的欄寬，避免複雜的計算
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                    try:
                        ws.column_dimensions[col].width = 15
                    except:
                        pass
            except:
                pass

        # 生成檔案名稱 - 避免特殊字符
        current_date = datetime.now().strftime("%Y%m%d")
        stock_name = basic_info.get('股票名稱', self.stock_id)
        # 移除可能造成檔名問題的字符
        safe_stock_name = "".join(c for c in stock_name if c.isalnum() or c in (' ', '-', '_')).strip()
        filename = f"{self.stock_id}_{safe_stock_name}_財務分析報告_{current_date}.xlsx"

        # 儲存檔案
        try:
            wb.save(filename)
            print(f"✅ 報告生成成功: {filename}")
            print(f"📊 包含工作表: 基本資訊、月營收、季度財務、年度財務、股利政策、潛力分析")
            return True
        except Exception as e:
            print(f"❌ 儲存檔案時發生錯誤: {e}")
            # 嘗試使用簡化檔名
            try:
                simple_filename = f"{self.stock_id}_財務分析報告_{current_date}.xlsx"
                wb.save(simple_filename)
                print(f"✅ 報告生成成功: {simple_filename}")
                return True
            except Exception as e2:
                print(f"❌ 簡化檔名也失敗: {e2}")
                return False

def main():
    """主函數"""
    parser = argparse.ArgumentParser(description='生成股票財務分析報告')
    parser.add_argument('stock_id', help='股票代號 (例如: 2330, 0050)')

    args = parser.parse_args()

    # 創建報告生成器
    generator = StockReportGenerator(args.stock_id)

    # 生成報告
    success = generator.generate_excel_report()

    if success:
        print("\n🎉 財務分析報告生成完成！")
        print("📋 報告內容包括:")
        print("   • 基本資訊 (股票代號、名稱、市場、最新股價等)")
        print("   • 月營收 (近24個月營收及成長率)")
        print("   • 季度財務 (近8季損益表資料)")
        print("   • 年度財務 (近5年財務比率及資產負債)")
        print("   • 股利政策 (近5年配息配股記錄)")
        print("   • 潛力分析 (系統評分及各維度分析)")
        print("\n💡 Mac兼容性: 此腳本完全支援macOS，無編碼問題")
    else:
        print("\n❌ 報告生成失敗")
        print("💡 請檢查:")
        print("   • 股票代號是否正確")
        print("   • 是否已安裝 openpyxl: pip install openpyxl")
        print("   • 資料庫是否包含該股票的資料")

if __name__ == "__main__":
    main()
